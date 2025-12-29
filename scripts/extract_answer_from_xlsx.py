#!/usr/bin/env python3
"""Extract answer text from xlsx files in a folder.

This script scans all .xlsx files under a given directory (non-recursive by
default), finds the column named "answer" (case-insensitive) in each worksheet,
and replaces each cell's value with the content between:

  </think><answer> ... </answer>

It writes results to a new file next to the original file and never modifies
the original.
"""

from __future__ import annotations

import argparse
import logging
import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


_ANSWER_HEADER = "answer"
_DEFAULT_HEADER_SEARCH_MAX_ROWS = 50
_DEFAULT_HEADER_SEARCH_MAX_COLS = 200

_ANSWER_PATTERN = re.compile(
    r"</think>\s*<answer>(?P<answer>.*?)</answer>",
    flags=re.DOTALL | re.IGNORECASE,
)


def _safe_output_path(input_path: Path, suffix: str) -> Path:
  """Return a non-existing output path in the same directory."""
  if input_path.suffix.lower() != ".xlsx":
    raise ValueError(f"Expected .xlsx file, got: {input_path}")

  base = input_path.with_suffix("")
  candidate = base.with_name(base.name + suffix).with_suffix(".xlsx")
  if not candidate.exists():
    return candidate

  for i in range(1, 10_000):
    candidate = base.with_name(f"{base.name}{suffix}.{i}").with_suffix(".xlsx")
    if not candidate.exists():
      return candidate

  raise RuntimeError(f"Failed to find available output name for {input_path}")


def _find_answer_column(
    ws,
    max_header_rows: int,
    max_header_cols: int,
) -> Optional[tuple[int, int]]:
  """Find the (header_row, answer_col) for a worksheet.

  Returns:
    A tuple of (header_row_index_1based, answer_col_index_1based) if found,
    otherwise None.
  """
  max_row = min(ws.max_row or 0, max_header_rows)
  max_col = min(ws.max_column or 0, max_header_cols)
  if max_row <= 0 or max_col <= 0:
    return None

  for row in range(1, max_row + 1):
    for col in range(1, max_col + 1):
      value = ws.cell(row=row, column=col).value
      if value is None:
        continue
      if isinstance(value, str) and value.strip().lower() == _ANSWER_HEADER:
        return row, col
  return None


def _extract_answer_text(text: str) -> Optional[str]:
  """Extract answer text using the required tag pattern."""
  match = _ANSWER_PATTERN.search(text)
  if not match:
    return None
  return match.group("answer").strip()


def process_xlsx_file(
    xlsx_path: Path,
    output_suffix: str,
    max_header_rows: int,
    max_header_cols: int,
) -> Path:
  """Process one .xlsx file and write to a new file.

  Returns:
    Output file path.
  """
  wb = load_workbook(filename=str(xlsx_path))

  total_cells_considered = 0
  total_cells_matched = 0

  for ws in wb.worksheets:
    header = _find_answer_column(
        ws,
        max_header_rows=max_header_rows,
        max_header_cols=max_header_cols,
    )
    if header is None:
      continue
    header_row, answer_col = header

    for row in range(header_row + 1, (ws.max_row or header_row) + 1):
      cell = ws.cell(row=row, column=answer_col)
      if cell.value is None:
        continue
      if not isinstance(cell.value, str):
        continue

      text = cell.value
      if not text.strip():
        continue

      total_cells_considered += 1
      extracted = _extract_answer_text(text)
      if extracted is None:
        logging.warning(
            "Pattern not found: file=%s sheet=%s row=%d col=%d",
            str(xlsx_path),
            ws.title,
            row,
            answer_col,
        )
        continue

      cell.value = extracted
      total_cells_matched += 1

  if total_cells_considered == 0:
    logging.warning(
        "No non-empty string cells found under 'answer' column: file=%s",
        str(xlsx_path),
    )
  elif total_cells_matched == 0:
    logging.warning(
        "No cells matched the pattern in file=%s (considered=%d)",
        str(xlsx_path),
        total_cells_considered,
    )

  out_path = _safe_output_path(xlsx_path, output_suffix)
  wb.save(filename=str(out_path))
  return out_path


def _iter_xlsx_files(folder: Path, recursive: bool) -> list[Path]:
  pattern = "**/*.xlsx" if recursive else "*.xlsx"
  return sorted([p for p in folder.glob(pattern) if p.is_file()])


def main() -> int:
  parser = argparse.ArgumentParser(
      description=(
          "Extract text between </think><answer> and </answer> for the 'answer' "
          "column in all .xlsx files within a folder. Writes new files next to "
          "the originals."
      )
  )
  parser.add_argument(
      "folder",
      type=str,
      help="Input folder containing one or more .xlsx files.",
  )
  parser.add_argument(
      "--recursive",
      action="store_true",
      help="Also scan subfolders for .xlsx files.",
  )
  parser.add_argument(
      "--output-suffix",
      type=str,
      default="_extracted",
      help="Suffix added to output filename (before .xlsx).",
  )
  parser.add_argument(
      "--max-header-rows",
      type=int,
      default=_DEFAULT_HEADER_SEARCH_MAX_ROWS,
      help="Max rows to search for the 'answer' header.",
  )
  parser.add_argument(
      "--max-header-cols",
      type=int,
      default=_DEFAULT_HEADER_SEARCH_MAX_COLS,
      help="Max columns to search for the 'answer' header.",
  )
  parser.add_argument(
      "--log-level",
      type=str,
      default="INFO",
      choices=["DEBUG", "INFO", "WARNING", "ERROR"],
      help="Logging level.",
  )
  args = parser.parse_args()

  logging.basicConfig(
      level=getattr(logging, args.log_level),
      format="%(levelname)s: %(message)s",
  )

  folder = Path(args.folder).expanduser().resolve()
  if not folder.exists() or not folder.is_dir():
    logging.error("Folder not found or not a directory: %s", str(folder))
    return 2

  xlsx_files = _iter_xlsx_files(folder, recursive=args.recursive)
  if not xlsx_files:
    logging.warning("No .xlsx files found in folder: %s", str(folder))
    return 0

  for xlsx_path in xlsx_files:
    try:
      out_path = process_xlsx_file(
          xlsx_path=xlsx_path,
          output_suffix=args.output_suffix,
          max_header_rows=args.max_header_rows,
          max_header_cols=args.max_header_cols,
      )
      logging.info("Wrote: %s", str(out_path))
    except Exception as exc:  # pylint: disable=broad-exception-caught
      logging.error("Failed processing %s: %s", str(xlsx_path), str(exc))
      return 1

  return 0


if __name__ == "__main__":
  raise SystemExit(main())

